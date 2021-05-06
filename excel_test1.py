from openpyxl import load_workbook



workbook = load_workbook(filename="sample.xlsx")
print(workbook.sheetnames)

sheet = workbook.active
sheet
sheet.title

print(sheet["A1"])
print(sheet["A1"].value)
print(sheet["F10"].value)
print(sheet.cell(row=10, column=6))
print(sheet.cell(row=10, column=6).value)
print(sheet["A1:C2"])
print(len(sheet["A"]))

for row in sheet.iter_rows(min_row=1,max_row=1,values_only=True):
    print(row)

# for row in sheet.rows:
#     print(row.)