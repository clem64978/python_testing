"""Get data from excel module based on openpyxl library.

The functions and classes in this module retrieve data 
from excel files and serialize them based on the design output format.  

Classes:
    * create - functions to create spreadsheets based on a template
    * get - functions to get data out of a spreasheet
    * update - functions to update data in a spreadsheet
    * extract - functions to extract data from a spreadsheet
"""

import os
import json

from openpyxl import load_workbook

class SheetTable():
    """Get data from excel module based on openpyxl library."""

    products = {}
    workbook = workbookDesc()

    def __init__(self,fn):
        self.workbook = load_workbook(filename=fn)
        self.sheet = self.workbook.active

        # if sn = None:
        #     self.sheet = self.workbook.active
        # elif:
        #     self.sheet = self.workbook[sn].active
    
    def extract_data(self,sheet_name=None,serialization_type='JSON'):
        """Extracts spreadsheet data and serialize based on the specified type.

        Parameters
        ----------
        sheet_name : str
            The spreadsheets name within the selected worksheet
        serialization_type : str, optional
            serialization formats are CSV, JSON, XML, YAML

        Returns
        -------
        object
            a seriaized object based based on the specified type.
        """
        
        for row in self.sheet.iter_rows(values_only=True):
            # print(row)
            product_id = row[3]
            product = {
                "parent": row[4],
                "title": row[5],
                "category": row[6]
            }
            self.products[product_id] = product
        
        return self.products